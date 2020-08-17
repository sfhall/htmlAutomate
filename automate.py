import openpyxl

def abbreviate(state):
    stateAbbrev = {
        'Alabama': 'AL',
        'Alaska': 'AK',
        'American Samoa': 'AS',
        'Arizona': 'AZ',
        'Arkansas': 'AR',
        'California': 'CA',
        'Colorado': 'CO',
        'Connecticut': 'CT',
        'Delaware': 'DE',
        'District of Columbia': 'DC',
        'Florida': 'FL',
        'Georgia': 'GA',
        'Guam': 'GU',
        'Hawaii': 'HI',
        'Idaho': 'ID',
        'Illinois': 'IL',
        'Indiana': 'IN',
        'Iowa': 'IA',
        'Kansas': 'KS',
        'Kentucky': 'KY',
        'Louisiana': 'LA',
        'Maine': 'ME',
        'Maryland': 'MD',
        'Massachusetts': 'MA',
        'Michigan': 'MI',
        'Minnesota': 'MN',
        'Mississippi': 'MS',
        'Missouri': 'MO',
        'Montana': 'MT',
        'Nebraska': 'NE',
        'Nevada': 'NV',
        'New Hampshire': 'NH',
        'New Jersey': 'NJ',
        'New Mexico': 'NM',
        'New York': 'NY',
        'North Carolina': 'NC',
        'North Dakota': 'ND',
        'Northern Mariana Islands':'MP',
        'Ohio': 'OH',
        'Oklahoma': 'OK',
        'Oregon': 'OR',
        'Pennsylvania': 'PA',
        'Puerto Rico': 'PR',
        'Rhode Island': 'RI',
        'South Carolina': 'SC',
        'South Dakota': 'SD',
        'Tennessee': 'TN',
        'Texas': 'TX',
        'Utah': 'UT',
        'Vermont': 'VT',
        'Virgin Islands': 'VI',
        'Virginia': 'VA',
        'Washington': 'WA',
        'West Virginia': 'WV',
        'Wisconsin': 'WI',
        'Wyoming': 'WY'
    }

    return stateAbbrev.get(state, state) #if the state is found, return the abbrevation, else just return the state


#structure of excel file must be: title, company, location-city, location-state
filename = input("Input the excel sheet:") #gets excel filename
wb = openpyxl.load_workbook(filename) #open excel file
ws = wb.active #get active sheet
outputFile = open("placements.txt", "x") #creates new file for output

numOfRows = ws.max_row + 1

for x in range(1, numOfRows):
    state = abbreviate(ws.cell(row = x, column = 4).value)

    outputFile.write("<tr> \n")
    outputFile.write("\t<td>" + ws.cell(row = x, column = 1).value + "</td> \n")
    outputFile.write("\t<td>" + ws.cell(row = x, column = 2).value + "</td> \n")
    outputFile.write("\t<td>" + ws.cell(row = x, column = 3).value + ", " + state + "</td>\n")
    outputFile.write("</td> \n")
